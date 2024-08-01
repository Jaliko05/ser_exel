from openpyxl import load_workbook
from openpyxl import Workbook
import random

def buscar_var_en_xlsx(ruta_archivo):
    libro = load_workbook(ruta_archivo, data_only=True)
    resultado = {}

    for nombre_hoja in libro.sheetnames:
        hoja = libro[nombre_hoja]
        resultado[nombre_hoja] = {}

        for fila in hoja.iter_rows():
            for celda in fila:
                valor_celda = celda.value
                if isinstance(valor_celda, str) and valor_celda.startswith("<VAR"):
                    posicion_celda = f"{celda.row},{celda.column}"
                    resultado[nombre_hoja][valor_celda] = posicion_celda

    return resultado

def procesar_txt(ruta_archivo):
    resultado = {}
    subclave_actual = None

    with open(ruta_archivo, 'r', encoding='utf-8') as archivo:
        lineas = archivo.readlines()

        # Omitir la primera línea
        lineas = lineas[1:]

        for linea in lineas:
            partes = linea.split('')
            if not partes or len(partes) < 2:
                continue

            hoja = partes[0]
            posible_subclave = partes[1]

            if posible_subclave.isdigit() and len(posible_subclave) == 3:
                subclave_actual = posible_subclave
                if hoja not in resultado:
                    resultado[hoja] = {}
                if subclave_actual not in resultado[hoja]:
                    resultado[hoja][subclave_actual] = {}

                for i, valor in enumerate(partes[2:], start=1):
                    var_key = f"<VAR{str(i).zfill(3)}>"
                    resultado[hoja][subclave_actual][var_key] = valor

            else:
                if hoja not in resultado:
                    resultado[hoja] = {}

                subclave_actual = None

                for i, valor in enumerate(partes[1:], start=1):
                    var_key = f"<VAR{str(i).zfill(3)}>"
                    resultado[hoja][var_key] = valor

    return resultado



def actualizar_excel(ruta_excel, posiciones, datos_txt):
    libro = load_workbook(ruta_excel)
    wb_copy = libro

    for hoja, vars_posiciones in posiciones.items():
        if hoja in datos_txt:
            hoja_excel = wb_copy[hoja]
            for subclave, vars_valores in datos_txt[hoja].items():
                if isinstance(vars_valores, dict):
                    for var_key, valor in vars_valores.items():
                        if var_key in vars_posiciones:
                            fila_base, columna = map(int, vars_posiciones[var_key].split(','))
                            fila = fila_base + int(subclave) - 1  # Ajuste de índice para fila
                            hoja_excel.cell(row=fila, column=columna, value=valor)
                else:
                    for var_key, valor in datos_txt[hoja].items():
                        if var_key in vars_posiciones:
                            fila, columna = map(int, vars_posiciones[var_key].split(','))
                            hoja_excel.cell(row=fila, column=columna, value=valor)

    randon = random.randint(0, 1000000)
    nameArchive = ruta_excel.split('.')
    newName = nameArchive[0]+'_'+str(randon)+'.xlsx'
    wb_copy.save(newName)

ruta_archivo_txt = '000015656.txt'  # Cambia esto por la ruta correcta
datos_txt = procesar_txt(ruta_archivo_txt)

ruta_archivo_xlsx = 'FINE088.xlsx'  # Cambia esto por la ruta correcta
posiciones_excel = buscar_var_en_xlsx(ruta_archivo_xlsx)

actualizar_excel(ruta_archivo_xlsx, posiciones_excel, datos_txt)
