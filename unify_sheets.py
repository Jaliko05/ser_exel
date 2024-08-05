from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

def unify_sheets(input_file):
    # Load the input workbook
    workbook = load_workbook(input_file)
    if "PRINCIPAL" in workbook.sheetnames:
        principal_sheet = workbook["PRINCIPAL"]
    elif  "principal" in workbook.sheetnames:
        principal_sheet = workbook["principal"]
    elif  "Principal" in workbook.sheetnames:
        principal_sheet = workbook["Principal"]
    else:
        principal_sheet = workbook.create_sheet(title="PRINCIPAL")

    row_offset = 0

    for sheet_name in workbook.sheetnames:
        if sheet_name.upper() == "PRINCIPAL" or sheet_name == "principal" or sheet_name == "Principal":
            continue
        sheet = workbook[sheet_name]

        # Copy merged cells
        merged_cells_ranges = list(sheet.merged_cells.ranges)
        for merged_cell_range in merged_cells_ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_cell_range))
            min_row += row_offset
            max_row += row_offset
            principal_sheet.merge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)

        # Copy cell values and styles
        for row in sheet.iter_rows():
            for cell in row:
                new_cell = principal_sheet.cell(row=cell.row + row_offset , column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell._style = cell._style

            # Check for the "??FIN??" marker to determine row offset
            if any(cell.value == "??FIN??" for cell in row):
                row_offset = principal_sheet.max_row - 1

    # Save the workbook back to the original file
    workbook.save(input_file)

# Ejemplo de uso:
# input_file = 'PSRH2060_178663.xlsx'
# unify_sheets(input_file)
