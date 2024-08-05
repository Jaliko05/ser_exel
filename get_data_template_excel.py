from openpyxl import load_workbook

def get_data_template_excel(ruta_archivo):
    libro = load_workbook(ruta_archivo, data_only=True)
    resultado = {}
    posiciones_estilos = {}
    posiciones_fin = {}

    for nombre_hoja in libro.sheetnames:
        hoja = libro[nombre_hoja]
        resultado[nombre_hoja] = {}
        posiciones_estilos[nombre_hoja] = {}

        for fila in hoja.iter_rows():
            for celda in fila:
                valor_celda = celda.value
                if isinstance(valor_celda, str):
                    if valor_celda.startswith("<VAR"):
                        posicion_celda = f"{celda.row},{celda.column}"
                        resultado[nombre_hoja][valor_celda] = posicion_celda
                        estilo_celda = {
                            'font': celda.font,
                            'fill': celda.fill,
                            'border': celda.border,
                            'alignment': celda.alignment,
                            'number_format': celda.number_format,
                            'protection': celda.protection
                        }
                        posiciones_estilos[nombre_hoja][posicion_celda] = estilo_celda
                    elif valor_celda == "??FIN??":
                        posicion_fin = f"{celda.row},{celda.column}"
                        posiciones_fin[nombre_hoja] = posicion_fin
    
    return resultado, posiciones_estilos, posiciones_fin