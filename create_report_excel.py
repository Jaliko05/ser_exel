import random
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Alignment, Protection

def create_report_excel(ruta_excel, posiciones, posiciones_estilos, posiciones_fin, datos_txt):
    libro = load_workbook(ruta_excel)
    wb_copy = libro

    for hoja, vars_posiciones in posiciones.items():
        if hoja in datos_txt:
            hoja_excel = wb_copy[hoja]
            fila_fin, columna_fin = map(int, posiciones_fin[hoja].split(','))
            print(fila_fin, columna_fin)

            for subclave, vars_valores in datos_txt[hoja].items():
                for var_key, valor in vars_valores.items():
                    if var_key in vars_posiciones:
                        fila_base, columna = map(int, vars_posiciones[var_key].split(','))
                        fila = fila_base + int(subclave) - 1 
                        celda_objetivo = hoja_excel.cell(row=fila, column=columna, value=valor)
                        
                        # Aplicar los estilos
                        posicion_celda = f"{fila_base},{columna}"
                        if hoja in posiciones_estilos and posicion_celda in posiciones_estilos[hoja]:
                            estilo_celda = posiciones_estilos[hoja][posicion_celda]
                            celda_objetivo.font = Font(
                                name=estilo_celda['font'].name,
                                size=estilo_celda['font'].size,
                                bold=estilo_celda['font'].bold,
                                italic=estilo_celda['font'].italic,
                                vertAlign=estilo_celda['font'].vertAlign,
                                underline=estilo_celda['font'].underline,
                                strike=estilo_celda['font'].strike,
                                color=estilo_celda['font'].color
                            )
                            celda_objetivo.fill = PatternFill(
                                fill_type=estilo_celda['fill'].fill_type,
                                start_color=estilo_celda['fill'].start_color,
                                end_color=estilo_celda['fill'].end_color
                            )
                            celda_objetivo.border = Border(
                                left=estilo_celda['border'].left,
                                right=estilo_celda['border'].right,
                                top=estilo_celda['border'].top,
                                bottom=estilo_celda['border'].bottom,
                                diagonal=estilo_celda['border'].diagonal,
                                diagonal_direction=estilo_celda['border'].diagonal_direction,
                                outline=estilo_celda['border'].outline,
                                vertical=estilo_celda['border'].vertical,
                                horizontal=estilo_celda['border'].horizontal
                            )
                            celda_objetivo.alignment = Alignment(
                                horizontal=estilo_celda['alignment'].horizontal,
                                vertical=estilo_celda['alignment'].vertical,
                                text_rotation=estilo_celda['alignment'].text_rotation,
                                wrap_text=estilo_celda['alignment'].wrap_text,
                                shrink_to_fit=estilo_celda['alignment'].shrink_to_fit,
                                indent=estilo_celda['alignment'].indent
                            )
                            celda_objetivo.number_format = estilo_celda['number_format']

            # Volver a escribir ??FIN?? si había múltiples repeticiones
            if len(datos_txt[hoja]) > 1:
                if hoja != 'Principal':
                    # Eliminar ??FIN?? si son rep
                    for row in hoja_excel.iter_rows():
                        for cell in row:
                            if cell.value == "??FIN??":
                                cell.value = None
                # Volver a escribir ??FIN?? en la ultima fila de la rep
                cell = hoja_excel.cell(row=fila + 1, column=columna_fin, value="??FIN??")
                cell.font = Font(color="FFFFFF")

    randon = random.randint(0, 1000000)
    nameArchive = os.path.splitext(ruta_excel)[0]
    newName = nameArchive + '_' + str(randon) + '.xlsx'   
    wb_copy.save(newName)
    return newName

# Ejemplo de uso:
# resultado, posiciones_estilos, posiciones_fin = get_data_template_excel('PSRH2060.xlsx')
# datos_txt = {
#     'Sheet1': {
#         '1': {
#             '<VAR1>': 'Nuevo valor 1',
#             '<VAR2>': 'Nuevo valor 2'
#         }
#     }
# }
# create_report_excel('PSRH2060.xlsx', resultado, posiciones_estilos, posiciones_fin, datos_txt)
